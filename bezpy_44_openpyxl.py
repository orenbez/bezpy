# Try this: https://towardsdatascience.com/automate-excel-reporting-with-python-233dd61fb0f2
# Try this: https://analyticsindiamag.com/guide-to-openpyxl-a-python-module-for-excel/

# more styles here https://openpyxl.readthedocs.io/en/stable/styles.html
# sys.getdefaultencoding() returns the default encoding
# if you get this error ... openpyxl.utils.exceptions.illegalcharactererror  then use the option

# xlrd  is the base library for reading data and formatting information in Excel files

# xlwt is another excel library for managing excel. It was designed to work with Excel files versions 95 through
# to 2003, which was the binary format prior to the OOXML (Open Office XML) format that was introduced with Excel 2007.

# xlutils is a continuation of xlrd and xlwt.
# The package provides more extensive setr of APIs for working with xls based Excel files
# in panda's ExcelWriter or to_excel() function use engine=xlwt

# openpyxl & xlswriter support the OOXML format Excel, which means 2007 onwards.
# in panda's ExcelWriter or to_excel() function use engine=openpyxl, engine=xlswriter or engine=xlwt

# openpylx is a library to read and write Excel 2010 files. You can write new worksheets, edit existing ones,
# and do pretty much any operation you can do using the mouse in Excel.
# It supports pretty much any Excel extension! and it is integrated with Pandas

# xlsxwriter library is better for applying styles to a range,  It is a full feature package including formatting,
# cell manipulation, formulas, pivot tables, charts, filters, data validation and drop-down list, memory optimization
# and images, conditional formatting, merged cells. it is integrated with Pandas

# pywin32 not specifically for Excel. Rather, it is a Python wrapper for the Windows API which provides access to COM
# (Common Object Model). COM is a common interface to all Windows based applications, Microsoft Office including Excel

# mito is another spreadsheet interface for python https://youtu.be/eF2QV4ymapk


import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Color, Fill, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# ===================================
# EXAMPLE 1: Format existing workbook
# ===================================

file = r'.\myfiles\PythonExport.xlsx'
wb = load_workbook(file) # Existing Workbook
sheet = wb['Sheet5']
read_value = sheet['F5'].value            # Method 1
read_value = sheet.cell(row=3,column=2)   # Method 2
sheet['F9'] = '=SUM(F2:F8)'
sheet['F9'].style = 'Currency'


max_row=sheet.max_row # get max row count
max_column=sheet.max_column # get max column count


wb.save(file)


# ===================================
# EXAMPLE 2
# ===================================
df = pd.read_excel('https://github.com/datagy/mediumdata/raw/master/sample_pivot.xlsx', parse_dates=['Date'])
print(df.head())


# Testing Pivot Tables
filtered = df[df['Region'] == 'East']
quarterly_sales = pd.pivot_table(filtered, index=filtered['Date'].dt.quarter, columns='Type', values='Sales', aggfunc='sum')

print("Quarterly Sales Pivot Table:")
print(quarterly_sales.head())

file_path = r'.\myfiles\PythonExport3.xlsx'
quarterly_sales.to_excel(file_path, sheet_name = 'Quarterly Sales', startrow=3)

wb = load_workbook(file_path) # Loading Existing Workbook
sheet1 = wb['Quarterly Sales']

# Formatting the First Sheet
sheet1['A1'] = 'Quarterly Sales'
sheet1['A2'] = 'datagy.io'
sheet1['A4'] = 'Quarter'

sheet1['A1'].style = 'Title'
sheet1['A2'].style = 'Headline 2'

for i in range(5, 9):
    sheet1[f'B{i}'].style='Currency'
    sheet1[f'C{i}'].style='Currency'
    sheet1[f'D{i}'].style='Currency'

# Adding a Bar Chart
bar_chart = BarChart()
data = Reference(sheet1, min_col=2, max_col=4, min_row=4, max_row=8)
categories = Reference(sheet1, min_col=1, max_col=1, min_row=5, max_row=8)
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)
sheet1.add_chart(bar_chart, "F4")
bar_chart.title = 'Sales by Type'
bar_chart.style = 3

wb.save(filename=file_path)

# ===============================
# EXAMPLE 3: Create New Workbook
# ===============================

wb = Workbook()          # Create New Workbook
ws = wb.active           # active worksheet
# wb.get_sheet_by_name('Sheet')      # DEPRECATED
ws = wb['Sheet']                     # same as above
ws = wb.worksheets[0]                # Access first worksheet
ws.title = 'sheet_name'
wb.remove(ws)                        # Delete the active sheet

ws = wb.create_sheet(title="NewSheet")  # Create a new sheet

# Assign Values to top Row
ws.cell(1, 1, "COL1 - VERY LONG NAME")   # (row,column, value)
ws.cell(1, 2, "COL2")
ws.cell(1, 3, "COL3")

# Alternatively way to assign Values for 2nd row
ws['A2'] = 5
ws['B2'] = 6
ws['C2'] = 7


# Alternatively way to assign Values for 3rd row
ws.append([8,9,10])

ws['C4'] = '=SUM(C2:C3)'  # Add a formula to a cell

# Note: can only apply style to a cell, row or column but not to a range
top_row = ws.row_dimensions[1]
top_row.font = Font(name='Calibri',  # DOESNT SEEM TO WORK
                    size=14,
                    bold=True,
                    italic=True,
                    vertAlign=None,
                    underline=None,
                    strike=False,
                    color="00FFFFFF")



bgcolor = PatternFill(fgColor="6e0c16", fill_type = "solid")

alignment=Alignment(horizontal='general',
                    vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                    shrink_to_fit=False,
                    indent=0)
bd = Border(outline=Side(border_style=None,color='FF000000'))

ws['A1'].fill = bgcolor
ws['B1'].fill = bgcolor
ws['C1'].fill = bgcolor

ws['A1'].font = Font(bold=True, color="FFFFFF")
ws['B1'].font = Font(bold=True, color="FFFFFF")
ws['C1'].font = Font(bold=True, color="FFFFFF")


# For a range you can use  ...
#
# my_header = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']
# for cell in my_header:



ws.column_dimensions['A'].auto_size = True   # DOESNT SEEM TO WORK
#ws.column_dimensions['A'].alignment.horizontal
ws.column_dimensions['B'].width = 30   # Set Explicit Size to Column 'B' width
ws.column_dimensions[get_column_letter(3)].auto_size = True  # This is column 'C'  # DOESNT SEEM TO WORK
ws.column_dimensions['B'].border = bd # doesn't work



for col in range(5,8):
    letter = get_column_letter(col)
    for row in range(1,4):
        ws[letter + str(row)] = letter + str(row)

# Delete a row
ws.delete_rows(7)

# Delete a column. Use the number of the col.
ws.delete_cols(8)

# Insert a row
ws.insert_rows(7)

# Insert a column. Use the number of the col.
ws.insert_cols(8)

wb.save(r'.\myfiles\PythonExport4.xlsx')


# ===============================
# Example 4
# ===============================
path = r'.\myfiles\PythonExport6.xlsx'
wb = load_workbook(path)
ws = wb.active # returns active sheet
ws.cell(column=1, row=ws.max_row + 1, value='TEST_1')  # Moves down to new row
ws.cell(column=2, row=ws.max_row, value='TEST_2')      # stays on same row
ws.cell(column=3, row=ws.max_row, value='TEST_3')

wb.save(path)